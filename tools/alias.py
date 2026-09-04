#!/usr/bin/env python3
"""
🥗 CONVERTIT LA TABLE D'ALIAS ALIMENTAIRES EN `data/alias.json`.

Michel : *« il faut que je puisse trouver les noms comme big Mac ou pizza 4 fromages »* pour le
fast-food (ft-v1114) — et, le même jour, pour les aliments courants : quelqu'un tape
« tortiglioni », « sticky rice » ou « riz jasmin », et l'app ne rend RIEN alors que l'aliment
générique est dans le téléphone.

La table vient de **GPT**, construite sur l'export CSV de notre propre base (le rapport d'audit du
03/09). ⭐⭐ ELLE NE CRÉE AUCUNE VALEUR NUTRITIONNELLE : un alias est une **porte** vers un code
CIQUAL existant, jamais une fiche. Vérifié ligne à ligne à l'import — voir les garde-fous.

⛔⛔ LES CINQ GARDE-FOUS, ET CHACUN A ATTRAPÉ QUELQUE CHOSE OU PEUT LE FAIRE :
  ① le code doit EXISTER dans `data/ciqual.json` ;
  ② les 4 macros doivent être IDENTIQUES aux nôtres — c'est ce qui prouve qu'aucune valeur n'a
     été réécrite en chemin (mesuré à l'import de la V2 : 0 écart sur 569 lignes) ;
  ③ ⚠️ LA CIBLE DOIT ÊTRE PROPOSABLE (kcal non `null`). **Ce contrôle a attrapé une vraie erreur** :
     `tomate` visait `20189 · Tomate, séchée`, dont les calories sont « non déterminées » dans
     CIQUAL — l'app ne peut JAMAIS la proposer. *Un alias qui ouvre une porte fermée est pire
     qu'un alias absent : la personne croit avoir cherché.*
  ④ un alias ne doit PAS être déjà dans `FOOD_SYNONYMES` (R2 : une information, un propriétaire).
     Les deux tables répondent à deux questions différentes — « traduire les mots tapés »
     (query → query, les familles et les marques) et « ce mot EST cet aliment » (query → code) —
     mais un même mot ne peut pas relever des deux ;
  ⑤ les lignes `HORS_CIQUAL` (sans code) sont ÉCARTÉES et LISTÉES : ce sont de vrais trous de la
     table nationale (whey, créatine, naan…), pas des erreurs. On les nomme au lieu de les taire.

⭐ POURQUOI UN CODE ET NON UNE REQUÊTE. `FOOD_SYNONYMES` (ft-v1113) traduit une frappe en une autre
frappe, puis laisse la recherche classer. Ça marche pour ouvrir une FAMILLE (« mcdo »), pas pour
désigner UN aliment : le classement trie par longueur de nom, et c'est ainsi que « steak haché
15 % » remonte du **veau** avant le **bœuf**. Un alias qui porte le CODE ne subit aucun classement.

⛔ LES CORRECTIONS APPORTÉES À LA TABLE sont ci-dessous, nommées, avec leur raison. Elles ne sont
pas silencieuses : le script les imprime à chaque exécution.

Usage : python3 tools/alias.py <fichier.xlsx>   → écrit data/alias.json
"""
import sys, json, re, os, unicodedata

# ⛔ CORRECTIONS EXPLICITES DE LA TABLE FOURNIE — chacune avec sa raison, jamais en silence.
#    On ne « répare » que ce qui est démontrablement faux, et on l'écrit (R30).
CORRECTIONS = {
    # `tomate` visait 20189 « Tomate, séchée » : kcal non déterminées → jamais proposable, et une
    # tomate n'est pas une tomate séchée. CIQUAL porte l'aliment moyen, c'est sa convention même
    # pour un mot générique — on ne choisit pas une variété à la place de la personne (R29).
    'tomate':  (20385, 'GPT visait « Tomate, séchée » (kcal non déterminées, jamais proposable)'),
    'tomates': (20385, 'idem'),
}

# ⛔ AJOUTS QUI NE VIENNENT PAS DU CLASSEUR — nommés, avec leur raison, jamais silencieux.
#    ⚠️ ILS VIVENT ICI ET NON DANS `data/alias.json` : ce fichier est GÉNÉRÉ, donc une retouche
#    à la main y disparaîtrait à la prochaine exécution, sans bruit (R27 : ce qui est généré ne
#    s'édite pas à la main).
AJOUTS = {
    # 🥤 LE CAS QUI LES A FAIT NAÎTRE (03/09/2026, ft-v1116) — signalé par Michel en notant son
    #    repas. `coca zéro` rendait « Cola, SUCRÉ, avec édulcorants » (18037, **24 kcal/100 g**)
    #    AVANT « Cola, SANS SUCRES AJOUTÉS, avec édulcorants » (18060, **1 kcal/100 g**).
    #    ⛔⛔ Sur une canette de 50 cl : **120 kcal enregistrées au lieu de 5**.
    #    👉 Cause : `zero`/`light` sont traduits en « édulcorants », et les DEUX lignes portent ce
    #    mot ; c'est alors le tri par NOM LE PLUS COURT qui tranche — et il tranche mal.
    #    *Un mot traduit désigne une famille, pas un aliment : quand la famille contient le
    #    contraire de ce qu'on cherche, il faut le code.*
    #    ⚠️ On ne retire PAS 18037 : c'est le bon aliment pour un cola sucré aux édulcorants
    #    (type stévia). On l'empêche seulement de répondre à la place du zéro.
    'coca zero':      (18060, 'rendait « Cola, sucré, avec édulcorants » (24 kcal/100 g) — 120 kcal au lieu de 5 sur 50 cl'),
    'coca light':     (18060, 'idem'),
    'coke zero':      (18060, 'idem'),
    'coke light':     (18060, 'idem'),
    'cola zero':      (18060, 'idem'),
    'cola light':     (18060, 'idem'),
    'coca sans sucre':(18060, 'rendait « Cola, sucré, SANS CAFÉINE » (41 kcal/100 g) — encore pire'),
    'cola sans sucre':(18060, 'idem'),
    'pepsi max':      (18060, 'même famille : un cola sans sucres ajoutés, aux édulcorants'),
    'pepsi zero':     (18060, 'idem'),
    'pepsi light':    (18060, 'idem'),
    # ⛔ Et la variante sans caféine a SA propre entrée dans la table nationale : on ne la
    #    confond pas avec la précédente, elles ne portent pas les mêmes valeurs.
    'coca zero sans cafeine': (18068, 'CIQUAL distingue la version sans caféine — on ne fusionne pas'),
    'coca sans cafeine':      (18067, 'le cola sucré sans caféine, qui existe aussi'),

    # ═══ 🥤 ft-v1117 — LE PIÈGE DU COCA N'ÉTAIT PAS CELUI DU COCA : IL Y EN A 9 ═══════════════
    #  Michel : *« et les autres boissons ? »*. **Mesuré** : CIQUAL porte **9 paires**
    #  « X, sucré, avec édulcorants » / « X, sans sucres ajoutés, avec édulcorants », et
    #  ⛔⛔ « sucré » est TOUJOURS le nom le plus court — donc le tri choisit systématiquement
    #  la version SUCRÉE quand on tape « light » ou « zéro ». Le Coca n'était qu'un cas sur 9.
    #  ⚠️⚠️ ET LE CORRECTIF « PROPRE » A ÉTÉ REFUSÉ PAR LA MESURE, pas par l'intuition :
    #  traduire `zero`/`light` en « sans sucres ajoutés » plutôt qu'en « édulcorants » corrige
    #  3 cas… et en casse 3 autres — `yaourt light` ne rend **plus rien**, `soda light` tombe
    #  sur « Boisson gazeuse **à la pomme** » (30 kcal, un autre produit).
    #  👉 ***« Light » ne veut pas dire « sans sucre » : un yaourt light est 0 % de matière
    #  grasse.*** Un mot qui a deux sens ne se traduit pas, il se DÉSIGNE — d'où ces alias.
    'soda zero':          (18001, 'les 9 paires : « sucré » est toujours le nom le plus court, donc il gagnait'),
    'soda light':         (18001, 'idem'),
    'soda sans sucre':    (18001, 'idem'),
    'boisson gazeuse zero':  (18001, 'idem'),
    'boisson gazeuse light': (18001, 'idem'),
    'limonade zero':      (18035, 'idem — 8 kcal/100 g au lieu de 0'),
    'limonade light':     (18035, 'idem'),
    'limonade sans sucre':(18035, 'idem'),
    'tonic zero':         (18013, 'idem — 22 kcal/100 g au lieu de 0'),
    'tonic light':        (18013, 'idem'),
    'schweppes zero':     (18013, 'idem'),
    'schweppes light':    (18013, 'idem'),
    'ice tea zero':       (18065, 'idem — 17 kcal/100 g au lieu de 1'),
    'ice tea light':      (18065, 'idem'),
    'the glace zero':     (18065, 'idem'),
    'the glace light':    (18065, 'idem'),
    'red bull zero':      (18353, 'idem — 28 kcal/100 g au lieu de 9'),
    'red bull sans sucre':(18353, 'idem'),
    'energy drink zero':  (18353, 'idem'),
    'boisson energisante zero': (18353, 'idem'),

    # ═══ 🥤 LES BOISSONS QUI NE RENDAIENT **RIEN** ══════════════════════════════════════════
    #  ⛔ Elles ouvrent le GÉNÉRIQUE de la table nationale, jamais un chiffre de marque : le nom
    #  affiché dit « Boisson énergisante » ou « Boisson gazeuse aux fruits », donc rien ne se
    #  fait passer pour le produit lui-même (R32/R33). C'est le mécanisme de « mcdo » en ft-v1113.
    #  ⭐ Et on prend l'**aliment moyen** quand il existe : c'est la convention de CIQUAL pour un
    #  mot générique, et elle ne choisit pas une variante à la place de la personne (R29).
    'ice tea':      (18062, 'ne rendait RIEN — l\'aliment moyen « Boisson au thé, aromatisée »'),
    'icetea':       (18062, 'idem'),
    'the glace':    (18062, 'idem'),
    'red bull':     (18324, 'ne rendait RIEN — l\'aliment moyen « Boisson énergisante »'),
    'redbull':      (18324, 'idem'),
    'monster':      (18324, 'idem'),
    'energy drink': (18324, 'idem'),
    'orangina':     (18048, 'ne rendait RIEN — l\'aliment moyen « Boisson gazeuse aux fruits, sucrée »'),
    'sprite':       (18048, 'idem'),
    'fanta':        (18048, 'idem'),
    'oasis':        (18048, 'idem'),
    'seven up':     (18048, 'idem'),
    'tropico':      (18048, 'idem'),
    'schweppes':    (18344, 'ne rendait RIEN — « Tonic ou bitter, sucré »'),
    'latte':        (18151, 'ne rendait RIEN — « Café au lait, café crème ou cappuccino, prêt à boire »'),
    'cafe latte':   (18151, 'idem'),
    'eau petillante': (18046, 'ne rendait RIEN — l\'aliment moyen « Eau minérale, gazeuse »'),
    'biere blonde': (5001,  'ne rendait RIEN — la bière standard « coeur de marché (4-5° alcool) »'),
    'panache':      (5004,  'ne rendait RIEN — « Panaché (limonade et bière) »'),

    # ═══ 🥛 LES LAITS VÉGÉTAUX — LE PIRE DE TOUS ═══════════════════════════════════════════
    #  ⛔⛔ `lait amande` rendait **« Chocolat au lait aux fruits secs (noisettes, amandes) » à
    #  559 kcal/100 g** — pour une boisson qui en fait **36**. ***× 15,5.*** Et `lait soja`,
    #  `lait avoine` ne rendaient RIEN. Cause : CIQUAL écrit « Boisson à… », on dit « lait de… ».
    #  ⚠️⚠️ ET ON NE TOUCHE PAS AU « LAIT DE COCO » : la table distingue le **lait de coco
    #  culinaire** (18041, **199 kcal**) de la **boisson à la noix de coco** (18907, **30**).
    #  Ce sont deux produits différents ; les fusionner ferait exactement le dégât qu'on répare.
    #  ⭐ Mesuré : « lait de coco » trouve DÉJÀ le bon (18041), il n'a besoin d'aucun alias.
    'lait amande':    (18111, 'rendait « Chocolat au lait aux fruits secs » à 559 kcal/100 g — × 15,5'),
    'lait d amande':  (18111, 'idem'),
    'lait damande':   (18111, 'idem'),
    'lait soja':      (18113, 'ne rendait RIEN — CIQUAL écrit « Boisson au soja »'),
    'lait de soja':   (18113, 'idem'),
    'lait avoine':    (18899, 'ne rendait RIEN — CIQUAL écrit « Boisson à l\'avoine »'),
    'lait d avoine':  (18899, 'idem'),
    'lait de riz':    (18904, 'ne rendait RIEN — CIQUAL écrit « Boisson au riz »'),

    # ⛔ « rosé » rendait « ROSETTE ou fuseau » — le SAUCISSON, à 392 kcal/100 g pour un vin qui
    #    en fait 69. C'est la famille de `hampe` → « cham**pêtre** » : la recherche compare des
    #    SOUS-CHAÎNES, donc « rose » se trouve dans « Rosette ».
    #    ⭐ Et rien n'est perdu : la rosette reste trouvable en tapant « rosette », et elle
    #    apparaît juste sous le vin dans la liste (R29).
    'rose':     (5216, 'rendait « Rosette ou fuseau » — le saucisson, 392 kcal au lieu de 69'),
    'vin rose': (5216, 'idem'),
    '7up':      (18048, 'ne rendait RIEN — l\'aliment moyen « Boisson gazeuse aux fruits, sucrée »'),
}

def norm(t):
    """La MÊME normalisation que `_afNorm` dans app.js — sinon la clé ne serait jamais trouvée."""
    t = str(t or '').lower()
    t = unicodedata.normalize('NFD', t)
    t = ''.join(c for c in t if unicodedata.category(c) != 'Mn')
    t = t.replace('œ', 'oe').replace('æ', 'ae')
    t = re.sub(r"['’‘`´ʼ]", '', t)
    # ⛔⛔ MIROIR DE ft-v1119 : la ponctuation devient un ESPACE, exactement comme dans `_afNorm`.
    #    ⛔ La barre `/` en est ABSENTE des deux côtés, exprès : la clé `lait 1/2 ecreme` la
    #    porte, et l'espacer d'un seul côté la rendrait introuvable.
    #    *Deux normalisations qui divergent d'un caractère font disparaître une entrée en silence.*
    t = re.sub(r'[,;:!?()\[\]{}«»"–—]+', ' ', t)   # ⛔ pas le « / » : voir _afNorm
    return re.sub(r'\s+', ' ', t).strip()

def synonymes_existants(app_js):
    """Les clés de FOOD_SYNONYMES, pour interdire qu'un mot relève des deux tables (R2)."""
    s = open(app_js, encoding='utf-8').read()
    i = s.find('const FOOD_SYNONYMES')
    if i < 0: return set()
    bloc = s[i:s.index('\n};', i)]
    bloc = re.sub(r'/\*.*?\*/', '', bloc, flags=re.S)
    return {norm(m) for m in re.findall(r"'([^']+)'\s*:", bloc)}

def main(src):
    import openpyxl
    ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    ciq = json.load(open(os.path.join(ROOT, 'data', 'ciqual.json'), encoding='utf-8'))
    base = {a[0]: a for a in ciq['a']}
    deja = synonymes_existants(os.path.join(ROOT, 'app.js'))

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    lignes = []
    for feuille in wb.sheetnames:
        ws = wb[feuille]
        it = ws.iter_rows(values_only=True)
        try: cols = next(it)
        except StopIteration: continue
        if not cols or 'alias' not in [str(c) for c in cols]: continue
        if 'code_ciqual' not in [str(c) for c in cols]: continue
        if 'libellé_CIQUAL' not in [str(c) for c in cols]: continue   # feuille d'import allégée
        for r in it:
            if r and r[0]: lignes.append(dict(zip(cols, r)))

    out, vus = {}, {}
    rej = {'hors_ciqual': [], 'inconnu': [], 'non_proposable': [], 'macro': [],
           'collision': [], 'doublon': [], 'court': []}
    corrigees, ajoutees = [], []

    for l in lignes:
        a = norm(l['alias'])
        if len(a) < 3:                    rej['court'].append(l['alias']); continue
        if a in deja:                     rej['collision'].append(l['alias']); continue

        code, raison = l.get('code_ciqual'), None
        if a in CORRECTIONS:
            code, raison = CORRECTIONS[a][0], CORRECTIONS[a][1]
        try: code = int(code)
        except (TypeError, ValueError):
            rej['hors_ciqual'].append((l['alias'], l.get('cible_affichée'))); continue

        e = base.get(code)
        if not e:                         rej['inconnu'].append((l['alias'], code)); continue
        # ③ la cible doit pouvoir être PROPOSÉE : sans calories, la ligne ne peut pas être notée.
        if e[3] is None:                  rej['non_proposable'].append((l['alias'], code, e[1])); continue

        # ② aucune valeur n'a été réécrite en chemin (sauté pour une correction : la ligne du
        #    classeur décrit alors un AUTRE aliment, c'est justement pour ça qu'on la corrige).
        if raison is None:
            ecart = False
            for cle, idx in (('kcal_100g', 3), ('prot_100g', 4), ('glucides_100g', 5), ('lipides_100g', 6)):
                v = l.get(cle)
                try: v = float(str(v).replace(',', '.'))
                except (TypeError, ValueError): v = None
                n = e[idx]
                if (v is None) != (n is None) or (v is not None and n is not None and abs(v - n) > 0.001):
                    ecart = True
            if ecart:                     rej['macro'].append((l['alias'], code)); continue
        else:
            corrigees.append((l['alias'], code, e[1], raison))

        if a in vus:
            # ⛔ deux cibles pour le même mot tapé = ambiguïté : on refuse les DEUX, on ne
            #    tranche pas au hasard (R29). Un doublon à l'identique, lui, est inoffensif.
            if vus[a] != code: rej['doublon'].append((l['alias'], vus[a], code)); out.pop(a, None)
            continue
        vus[a] = code
        out[a] = code

    # ⛔ LES AJOUTS PASSENT LES MÊMES GARDE-FOUS QUE LE CLASSEUR — sinon ils seraient une porte
    #    dérobée par laquelle une cible morte pourrait entrer.
    for a0, (code, raison) in AJOUTS.items():
        a = norm(a0)
        if a in deja:   rej['collision'].append(a0); continue
        e = base.get(code)
        if not e:       rej['inconnu'].append((a0, code)); continue
        if e[3] is None: rej['non_proposable'].append((a0, code, e[1])); continue
        if a in vus and vus[a] != code:
            rej['doublon'].append((a0, vus[a], code)); out.pop(a, None); continue
        vus[a] = code; out[a] = code
        ajoutees.append((a0, code, e[1], raison))

    dst = os.path.join(ROOT, 'data', 'alias.json')
    with open(dst, 'w', encoding='utf-8') as f:
        json.dump({'source': 'Alias FR → codes Ciqual 2025 (ANSES) — aucune valeur nutritionnelle créée',
                   'n': len(out), 'a': out}, f, ensure_ascii=False, separators=(',', ':'), sort_keys=True)

    ko = os.path.getsize(dst) / 1024
    import gzip as _gz
    kz = len(_gz.compress(open(dst, 'rb').read())) / 1024
    print(f'{len(out)} alias → {dst} ({ko:.0f} Ko · {kz:.0f} Ko gzippé)')
    if ajoutees:
        print(f'\n⭐ {len(ajoutees)} AJOUT(S) HORS CLASSEUR — nommés, avec leur raison :')
        for a, c, lib, r in ajoutees: print(f'   · « {a} » → {c} {lib}\n     raison : {r}')
    if corrigees:
        print(f'\n⛔ {len(corrigees)} CORRECTION(S) APPORTÉE(S) À LA TABLE — jamais en silence :')
        for a, c, lib, r in corrigees: print(f'   · « {a} » → {c} {lib}\n     raison : {r}')
    for cle, titre in (('hors_ciqual',   "écartés — ABSENTS de CIQUAL (vrais trous de la table nationale)"),
                       ('non_proposable', "écartés — cible sans calories déterminées (jamais proposable)"),
                       ('inconnu',       "écartés — code introuvable dans notre base"),
                       ('macro',         "écartés — macros divergentes (la valeur aurait été réécrite)"),
                       ('collision',     "écartés — déjà dans FOOD_SYNONYMES (R2, un propriétaire)"),
                       ('doublon',       "écartés — DEUX cibles pour le même mot (ambigu, on ne tranche pas)"),
                       ('court',         "écartés — moins de 3 caractères")):
        v = rej[cle]
        if not v: continue
        print(f'\n⚠️ {len(v)} {titre} :')
        for x in v[:14]: print('   ·', x)
        if len(v) > 14: print(f'   … et {len(v)-14} autres')

if __name__ == '__main__':
    main(sys.argv[1] if len(sys.argv) > 1 else 'Force_Tracker_Alias_CIQUAL_V2_Elargie.xlsx')
