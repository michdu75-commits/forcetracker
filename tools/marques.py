#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GÉNÉRATEUR DE `data/marques.json` — la base fast-food France (ft-v1114).

⭐⭐ POURQUOI UN SCRIPT ET PAS UNE SAISIE À LA MAIN (R27) : une table recopiée à la main devient
   fausse à la première mise à jour, et personne ne sait plus d'où vient un chiffre. Ici le
   classeur fourni par Michel est la SOURCE, ce fichier est la RECETTE, et `data/marques.json`
   est un produit — régénérable, donc vérifiable.

⛔⛔ CE QUE CE SCRIPT REFUSE D'IMPORTER, ET C'EST L'ESSENTIEL :

   ① UNE LIGNE INCOMPLÈTE. Il faut les QUATRE valeurs (kcal, protéines, glucides, lipides).
      Écrire 0 à la place d'une valeur absente serait inventer : une frite à 231 kcal avec
      « 0 g de glucides » fausserait les anneaux de la journée sans que rien ne le signale.
      C'est la règle de `tools/ciqual.py` : « - » veut dire NON DÉTERMINÉ, pas zéro.

   ② UNE LIGNE INCOHÉRENTE ENTRE TAILLES — et celle-là, aucun garde-fou de l'app ne peut la
      voir. Mesuré sur le classeur : les frites Quick donnent 588 kcal (petite), 594 (moyenne),
      596 (grande), pour une portion déduite IDENTIQUE de 181 g. Chaque ligne prise seule est
      parfaitement cohérente — les kcal collent aux macros (6×4 + 55×4 + 37×9 = 577 ≈ 588) et
      la matière tient dans la portion (98 g dans 181 g). *Le défaut n'existe qu'ENTRE les
      lignes.* Une petite frite doit peser moins qu'une grande : si ce n'est pas le cas, la
      famille entière est écartée plutôt qu'à moitié importée.

   ③ UNE LIGNE DONT LES kcal NE COLLENT PAS AUX MACROS de plus de 15 % ET de plus de 25 kcal.
      Les deux conditions ensemble, exprès : sur une dosette de vinaigrette à 11 kcal, 27 %
      d'écart font 3 kcal — c'est de l'arrondi, pas une erreur.

⛔ LA PROVENANCE VOYAGE AVEC LA VALEUR (R32/R33) : chaque aliment garde son enseigne, l'URL de
   la source officielle et la date de relevé. Rien ne se présente comme une mesure de l'app.

Lancer :  python3 tools/marques.py <classeur.xlsx>
"""
import sys, os, json, io, re
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl manquant :  pip install openpyxl")

RACINE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SORTIE = os.path.join(RACINE, 'data', 'marques.json')

# ⚠️ Les tailles, de la plus petite à la plus grande. Sert au contrôle ② : dans une même
#    famille de produit, une taille plus petite ne peut pas peser plus qu'une plus grande.
TAILLES = ['petite', 'petit', 'moyenne', 'moyen', 'grande', 'grand', 'maxi', 'xl']


def nombre(v):
    """Lit un nombre en acceptant la virgule française. Rend None si ce n'est pas un nombre —
       jamais 0, qui serait une valeur inventée."""
    if v is None or v == '':
        return None
    try:
        return float(str(v).replace(',', '.').strip())
    except ValueError:
        return None


def taille_de(nom):
    """Rend l'indice de taille trouvé dans le nom, ou None. Sert au contrôle entre tailles."""
    n = nom.lower()
    for i, t in enumerate(TAILLES):
        if re.search(r'\b' + t + r'\b', n):
            return i
    return None


def famille_de(nom):
    """Le nom débarrassé de sa taille : « Petite Frites » et « Grande Frites » sont la même
       famille, donc comparables entre elles."""
    n = nom.lower()
    for t in TAILLES:
        n = re.sub(r'\b' + t + r'\b', '', n)
    return re.sub(r'\s+', ' ', n).strip()


def main():
    if len(sys.argv) < 2:
        sys.exit("usage : python3 tools/marques.py <classeur.xlsx>")
    src = sys.argv[1]
    w = openpyxl.load_workbook(src, data_only=True)

    # ⚠️⚠️ ON NE LIT QU'UNE SEULE FEUILLE, ET C'EST UNE LEÇON PAYÉE : dans la version contrôlée,
    #    « FRITES vérifiées » est un SOUS-ENSEMBLE de « Base VALIDÉE » — les 14 lignes sont dans
    #    les deux. En concaténant les deux feuilles, mon contrôle entre tailles voyait
    #    « Petite 80 g < Petite 80 g » et déclarait la famille incohérente. *Un doublon
    #    ressemble exactement à une incohérence pour un détecteur qui compare des voisins.*
    feuille = next((n for n in ('Base VALIDÉE', 'Base validee', 'Base nutrition') if n in w.sheetnames), None)
    if not feuille:
        sys.exit("aucune feuille de référence trouvée (« Base VALIDÉE » ou « Base nutrition »)")
    print('feuille de référence :', feuille)
    s = w[feuille]
    hdr = [c.value for c in s[1]]
    brut = [dict(zip(hdr, r)) for r in s.iter_rows(min_row=2, values_only=True) if r[2] is not None]

    # ── les remarques par enseigne : elles portent les limites que la source déclare elle-même
    remarques = {}
    if 'Sources & couverture' in w.sheetnames:
        for r in w['Sources & couverture'].iter_rows(min_row=2, values_only=True):
            if r and r[0]:
                remarques[str(r[0])] = {'source': str(r[2] or ''), 'remarque': str(r[3] or '')}

    ecartes = defaultdict(list)
    lignes = []
    for x in brut:
        nom = str(x.get('Produit') or '').strip()
        ens = str(x.get('Enseigne') or '').strip()
        k = nombre(x.get('kcal portion'))
        p = nombre(x.get('Protéines g'))
        g = nombre(x.get('Glucides g'))
        l = nombre(x.get('Lipides g'))
        k100 = nombre(x.get('kcal /100g'))
        # ⚠️⚠️ J'AVAIS ÉCRIT L'INVERSE, ET LA MESURE M'A CORRIGÉE. J'avais transcrit les
        #   colonnes pour-100 g du classeur en me fiant à un contrôle trop large (« jamais plus
        #   de 1,5 g d'écart avec le dérivé »). Ce contrôle CACHAIT le défaut : les colonnes
        #   macro pour-100 g sont ARRONDIES À L'ENTIER par la source.
        #   Mesuré sur le Big Mac : 27 g de protéines pour 232 g font 11,64 g/100 g, que le
        #   classeur écrit « 12 ». Repartir de 12 rend 28 g sur la portion — soit 1 g de plus
        #   que le chiffre PUBLIÉ. Repartir de 11,6 rend 27 g, exactement le chiffre publié.
        #   👉 *Les valeurs par PORTION sont les chiffres primaires ; la colonne pour-100 g est
        #   un dérivé arrondi de la source.* On dérive donc du primaire, et on ne retombe sur la
        #   colonne que s'il n'y a pas de portion (cas Domino's).
        #   *Une tolérance trop large ne valide pas : elle dispense de regarder.*
        p100 = nombre(x.get('Protéines /100g'))
        g100 = nombre(x.get('Glucides /100g'))
        l100 = nombre(x.get('Lipides /100g'))

        # ── ① incomplète : on n'invente pas de zéro… MAIS ON SAIT CALCULER LES CALORIES.
        #    ⭐ Michel demande de pouvoir chercher « pizza 4 fromages » — c'était justement la
        #    ligne écartée : elle porte ses 3 macros et il ne manque QUE les kcal. Or les kcal
        #    se DÉDUISENT des macros (4/4/9), et l'app fait déjà ce calcul dans
        #    `_coherenceKcal`. Ce n'est donc pas une valeur inventée, c'est une valeur DÉRIVÉE —
        #    et elle est marquée comme telle (`kcal_derivee`), pour qu'on ne la présente jamais
        #    comme un chiffre publié (R32 : mesuré / estimé / propriétaire).
        #    ⛔ L'INVERSE RESTE REFUSÉ : des kcal sans macros ne se répartissent pas. On ne
        #    saurait pas quelle part est du gras et laquelle du sucre — c'est ça qui serait
        #    inventé (les frites McDo de l'ancien fichier, les 7 Subway).
        kcal_derivee = False
        if k is None and None not in (p, g, l):
            k = round(4 * p + 4 * g + 9 * l, 1)
            kcal_derivee = True
        if None in (k, p, g, l) or k <= 0:
            ecartes['incomplete'].append(ens + ' · ' + nom)
            continue

        # ── ③ les kcal collent-elles aux macros ? (les DEUX conditions, cf. en-tête)
        # ⛔⛔ ON NE JUGE PAS UNE VALEUR PAR LA FORMULE QUI VIENT DE LA PRODUIRE : un contrôle
        #    qui compare une kcal dérivée à 4P+4G+9L serait un vert qui ne peut pas rougir
        #    (`BUGS.md`, le corollaire de R33). On le saute donc pour ces lignes-là.
        theo = 4 * p + 4 * g + 9 * l
        if not kcal_derivee and abs(k - theo) > 25 and abs(k - theo) / k > 0.15:
            ecartes['kcal_incoherentes'].append('%s · %s (%d annoncées, %d calculées)' % (ens, nom, k, theo))
            continue

        # ── la portion se DÉDUIT du pour-100 g, elle ne s'invente pas
        # ⚠️ SANS pour-100 g kcal, la portion ne se déduit pas — mais si les valeurs SONT déjà
        #    pour 100 g (cas Domino's, la source le dit), l'aliment reste utilisable comme
        #    n'importe quel aliment de la table : 100 g par défaut, la personne met son poids.
        portion = round(k / k100 * 100) if k100 else None

        # ── la matière doit tenir dans la portion (règle physique de ft-v1103)
        if portion and (p + g + l) > portion + 2:
            ecartes['masse_impossible'].append('%s · %s (%.0f g dans %d g)' % (ens, nom, p + g + l, portion))
            continue

        lignes.append({
            'ens': ens, 'nom': nom, 'cat': str(x.get('Catégorie') or '').strip(),
            'kcal': round(k, 1), 'prot': round(p, 1), 'carbs': round(g, 1), 'fat': round(l, 1),
            # ⛔⛔ LES QUATRE se dérivent de la portion, kcal comprises — et ce n'est PAS
            #    circulaire, c'est une mise en cohérence. La portion (232 g) est déduite de la
            #    colonne kcal/100 g ARRONDIE du classeur (228), donc elle porte cet arrondi.
            #    Repartir de 228 rend 529 kcal sur 232 g ; le chiffre publié est 530. En
            #    dérivant 530/232 = 228,4, la portion redonne 530 exactement.
            #    👉 *Une fois la portion arrondie, la seule question est : quel pour-100 g
            #    reproduit le chiffre publié sur CETTE portion ?* Garder les deux arrondis en
            #    série faisait perdre 1 kcal et 1 g de protéines.
            'kcal100': round(k / portion * 100, 1) if portion else (round(k100, 1) if k100 else None),
            # ⛔ le pour-100 g ne se DÉDUIT qu'à défaut : la source d'abord.
            'prot100': round(p / portion * 100, 1) if portion else p100,
            'carbs100': round(g / portion * 100, 1) if portion else g100,
            'fat100': round(l / portion * 100, 1) if portion else l100,
            'portion': portion,
            'taille': taille_de(nom), 'famille': famille_de(nom),
            'kcal_derivee': kcal_derivee,
        })

    # ── ④ LE CONTRÔLE PAR PIÈCE : « 5 Tenders », « 7 Tenders + 7 Hot Wings », « 16 Tenders +
    #    16 Hot Wings » doivent donner un kcal PAR PIÈCE comparable. Mesuré sur le classeur
    #    contrôlé : 146, 77 et 38,5 kcal/pièce — *ça se divise par deux à chaque ligne*, donc
    #    les trois sont mutuellement incompatibles. La source signale le 16+16 comme douteux ;
    #    la mesure montre que les TROIS le sont. On les écarte toutes plutôt que d'en garder une
    #    au hasard : on ne sait pas laquelle est juste.
    PIECES = re.compile(r'\b(\d+)\s+(tenders?|hot\s*wings?|wings?|nuggets?|pi[eè]ces?|morceaux)\b', re.I)
    par_piece = defaultdict(list)
    for a in lignes:
        m = PIECES.findall(a['nom'])
        if not m:
            continue
        n = sum(int(q) for q, _ in m)
        if n > 0:
            a['_pieces'] = n
            par_piece[a['ens']].append(a)
    for ens, l in par_piece.items():
        if len(l) < 2:
            continue
        vals = [a['kcal'] / a['_pieces'] for a in l]
        if max(vals) > 2 * min(vals):     # un facteur 2 entre deux lignes de la même enseigne
            for a in l:
                a['_piece_ko'] = '%.0f kcal/pièce' % (a['kcal'] / a['_pieces'])

    # ── ② LE CONTRÔLE ENTRE TAILLES : dans une famille, plus petit doit peser moins.
    #    C'est le seul qui attrape les frites Quick, et il ne peut pas s'écrire ligne à ligne.
    par_famille = defaultdict(list)
    for a in lignes:
        if a['taille'] is not None and a['portion']:
            par_famille[(a['ens'], a['famille'])].append(a)
    familles_ko = set()
    for cle, fam in par_famille.items():
        if len(fam) < 2:
            continue
        fam.sort(key=lambda a: a['taille'])
        for i in range(1, len(fam)):
            # une taille plus grande qui ne pèse pas plus (à 5 g près) trahit la famille entière
            if fam[i]['portion'] <= fam[i - 1]['portion'] + 5:
                familles_ko.add(cle)
                break
    gardees = []
    for a in lignes:
        cle = (a['ens'], a['famille'])
        if cle in familles_ko:
            ecartes['tailles_incoherentes'].append('%s · %s (portion %s g)' % (a['ens'], a['nom'], a['portion']))
            continue
        if a.get('_piece_ko'):
            ecartes['par_piece_incoherent'].append('%s · %s (%s)' % (a['ens'], a['nom'], a['_piece_ko']))
            continue
        if a['kcal100'] is None or a['prot100'] is None or a['carbs100'] is None or a['fat100'] is None:
            # ⛔ SANS POUR-100 g COMPLET, le bloc quantité de l'app ne peut rien recalculer :
            #    l'aliment serait figé sur une seule portion. On l'écarte plutôt que de livrer
            #    une ligne qu'on ne peut pas ajuster (R29).
            ecartes['sans_pour_100g'].append(a['ens'] + ' · ' + a['nom'])
            continue
        gardees.append(a)

    # ── le fichier servi : compact, sans les champs de travail
    out = {
        'source': 'Base fast-food France (classeur fourni), sources officielles par enseigne',
        'genere_par': 'tools/marques.py',
        'note': ("Valeurs relevées sur les sources officielles citées par enseigne. L'application "
                 "ne les a PAS mesurées : elle les affiche en nommant leur origine. Les recettes "
                 "changent — la date de relevé fait partie de la donnee."),
        'enseignes': remarques,
        'champs': ['ens', 'nom', 'cat', 'kcal100', 'prot100', 'carbs100', 'fat100', 'portion', 'kcal_derivee'],
        'a': [[a['ens'], a['nom'], a['cat'], a['kcal100'], a['prot100'], a['carbs100'],
               a['fat100'], a['portion'], 1 if a['kcal_derivee'] else 0] for a in gardees],
    }
    os.makedirs(os.path.dirname(SORTIE), exist_ok=True)
    with io.open(SORTIE, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, separators=(',', ':'))

    print('%s : %d produits gardés sur %d lignes' % (os.path.relpath(SORTIE, RACINE), len(gardees), len(brut)))
    par_ens = defaultdict(int)
    for a in gardees:
        par_ens[a['ens']] += 1
    print('   par enseigne :', dict(par_ens))
    print('\n⛔ ÉCARTÉS, avec la raison (le silence serait le vrai défaut) :')
    for motif, l in sorted(ecartes.items()):
        print('   %-22s %3d' % (motif, len(l)))
        for x in l[:6]:
            print('        ·', x[:88])
        if len(l) > 6:
            print('        … et %d autres' % (len(l) - 6))


if __name__ == '__main__':
    main()
